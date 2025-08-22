import datetime
import time
from typing import Dict, List, Tuple, Optional
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ProductionResultProcessor:
    """
    생산 최적화 결과를 처리하고 출력하는 클래스
    """
    def __init__(self, optimizer):
        # optimizer 인스턴스 참조
        self.optimizer = optimizer
        
        # 필요한 데이터들을 optimizer에서 가져오기
        self.ontology_manager = optimizer.ontology_manager
        self.onto = optimizer.onto
        self.json_data = optimizer.json_data
        self.order_data = optimizer.order_data
        self.logger = optimizer.logger
        self.lines = optimizer.lines
        self.products = optimizer.products
        self.time_slots = optimizer.time_slots
        self.valid_product_line_combinations = optimizer.valid_product_line_combinations
        self.product_instances = optimizer.product_instances
        self.line_instances = optimizer.line_instances
        
        # 메서드들을 optimizer에서 가져오기
        self._get_product_name = optimizer._get_product_name
        self._get_capacity_rate = optimizer._get_capacity_rate
        self._get_track_count = optimizer._get_track_count
        self._get_package_count = optimizer._get_package_count
        self._get_changeover_time = optimizer._get_changeover_time
        self._get_setup_time = optimizer._get_setup_time
        self._get_max_working_hours = optimizer._get_max_working_hours

    def print_solution(self, solution: Dict):
        """
        최적화 결과 출력
        Args:
            solution: dict, 최적화 결과
        """
        if not solution:
            self.logger.error("출력할 결과가 없습니다")
            return
        
        print("\n" + "="*80)
        print("🏭 생산 계획 최적화 결과")
        print("="*80)
        
        # 제품별 목표 생산량 및 블록 분석 (신규 추가)
        self._print_product_analysis()
        
        # 목적함수 값
        print(f"목적함수 값: {solution['objective_value']:.2f}")
        
        # 통계 정보
        stats = solution['statistics']
        print(f"\n📊 통계 정보:")
        print(f"  총 생산시간: {stats['total_production_time']:.1f}시간")
        print(f"  총 교체시간: {stats['total_changeover_time']:.1f}시간")
        
        # 작업 준비 시간과 청소 시간 구분
        total_setup_time = 0
        total_cleanup_time = 0
        for event in solution['cleaning_events']:
            if event['time_slot'] == self.time_slots[0]:  # 첫 번째 시점 = 작업 준비 시간
                total_setup_time += event['cleaning_time']
            else:  # 나머지 시점 = 청소 시간
                total_cleanup_time += event['cleaning_time']
        
        print(f"  총 작업준비시간: {total_setup_time:.1f}시간")
        print(f"  총 청소시간: {total_cleanup_time:.1f}시간")
        print(f"  총 가동시간: {stats['total_working_time']:.1f}시간")
        
        # 교체시간 상세 분석
        print(f"\n🔄 품목교체시간 상세 분석:")
        print("-" * 80)
        
        # 라인별 교체시간 분석
        line_changeover_times = {}
        for event in solution['changeover_events']:
            line = event['line']
            if line not in line_changeover_times:
                line_changeover_times[line] = 0
            line_changeover_times[line] += event['changeover_time']
        
        if line_changeover_times:
            print(f"{'호기':<8} {'총 교체시간':<12} {'교체횟수':<8} {'평균교체시간':<12}")
            print("-" * 40)
            for line in sorted(line_changeover_times.keys()):
                total_time = line_changeover_times[line]
                count = len([e for e in solution['changeover_events'] if e['line'] == line])
                avg_time = total_time / count if count > 0 else 0
                print(f"{line:<8} {total_time:<12.1f} {count:<8} {avg_time:<12.1f}")
        else:
            print("  교체 이벤트가 없습니다.")
        
        # 시간대별 교체시간 분석
        print(f"\n⏰ 시간대별 교체시간 분석:")
        print("-" * 50)
        time_slot_changeover = {}
        for event in solution['changeover_events']:
            time_slot = event['time_slot']
            if time_slot not in time_slot_changeover:
                time_slot_changeover[time_slot] = 0
            time_slot_changeover[time_slot] += event['changeover_time']
        
        if time_slot_changeover:
            for time_slot in sorted(time_slot_changeover.keys()):
                total_time = time_slot_changeover[time_slot]
                count = len([e for e in solution['changeover_events'] if e['time_slot'] == time_slot])
                print(f"  {time_slot}: {total_time:.1f}시간 ({count}회 교체)")
        else:
            print("  교체 이벤트가 없습니다.")
        
        # 제품별 목표수량 vs 실제생산량 비교
        print(f"\n🎯 제품별 목표수량 vs 실제생산량 (박스 단위):")
        print("-" * 100)
        print(f"{'제품코드':<12} {'제품명':<25} {'목표수량':<10} {'실제생산량':<12} {'달성율':<8} {'생산호기':<15}")
        print("-" * 100)
        
        # 제품별 총 생산량 계산 (박스 단위)
        product_total_production = {}
        product_production_lines = {}  # 제품별 생산 호기 정보
        
        for line, schedule in solution['production_schedule'].items():
            for time_slot, productions in schedule.items():
                for prod in productions:
                    product_code = prod['product']
                    if product_code not in product_total_production:
                        product_total_production[product_code] = 0
                        product_production_lines[product_code] = set()
                    
                    product_total_production[product_code] += prod['production_quantity_boxes']
                    product_production_lines[product_code].add(line)
        
        # 제품별로 목표수량과 비교
        for product_code in sorted(self.products):
            target_quantity = self.order_data.get(product_code, 0)
            actual_quantity = product_total_production.get(product_code, 0)
            achievement_rate = (actual_quantity / target_quantity * 100) if target_quantity > 0 else 0
            
            # 제품명 가져오기
            product_name = self._get_product_name(product_code)
            
            # 생산 호기 정보
            production_lines = product_production_lines.get(product_code, set())
            lines_str = ", ".join(sorted(production_lines)) if production_lines else "미생산"
            
            # 달성율에 따른 색상 표시 (정확한 비교를 위해 소수점 처리)
            achievement_rate_rounded = round(achievement_rate, 1)
            if achievement_rate_rounded >= 100.0:
                achievement_status = "✅"
            elif achievement_rate_rounded >= 90.0:
                achievement_status = "⚠️"
            else:
                achievement_status = "❌"
            
            # 디버깅 정보 추가 (소수점 3자리까지 표시)
            print(f"{product_code:<12} {product_name:<25} {target_quantity:<10.0f} {actual_quantity:<12.0f} {achievement_rate_rounded:<7.1f}% {achievement_status} {lines_str:<15}")
            
            # 디버깅: 정확한 값 확인
            if abs(achievement_rate - 100.0) < 0.1 and achievement_status != "✅":
                print(f"    [디버깅] {product_code}: achievement_rate={achievement_rate:.6f}, rounded={achievement_rate_rounded:.1f}")
        
        print("-" * 100)
        
        # 전체 달성율 계산
        total_target = sum(self.order_data.values())
        total_actual = sum(product_total_production.values())
        total_achievement = (total_actual / total_target * 100) if total_target > 0 else 0
        print(f"전체 달성율: {total_achievement:.1f}% ({total_actual:.0f}박스/{total_target:.0f}박스)")
        
        # 생산 스케줄 (제품명 포함, 박스 단위) + 시간 분석
        print(f"\n📅 생산 스케줄 (박스 단위) + 시간 분석:")
        for line, schedule in solution['production_schedule'].items():
            print(f"\n  {line}호기:")
            for time_slot, productions in schedule.items():
                # 해당 시간대의 총 시간 계산
                total_production_time = sum(prod['production_time'] for prod in productions)
                
                # 교체시간 계산
                changeover_time = 0
                for event in solution['changeover_events']:
                    if event['line'] == line and event['time_slot'] == time_slot:
                        changeover_time += event['changeover_time']
                
                # 청소시간 계산 (작업 준비 시간과 청소 시간 구분)
                setup_time = 0
                cleanup_time = 0
                for event in solution['cleaning_events']:
                    if event['line'] == line and event['time_slot'] == time_slot:
                        if time_slot == self.time_slots[0]:  # 첫 번째 시점 (월요일 조간) = 작업 준비 시간
                            setup_time += event['cleaning_time']
                        elif time_slot == self.time_slots[-1]:  # 마지막 시점 (금요일 야간) = 청소 시간
                            cleanup_time += event['cleaning_time']
                        else:
                            cleanup_time += event['cleaning_time']  # 기타 시점은 청소 시간으로 처리
                
                # 총 시간 계산
                total_time = total_production_time + changeover_time + setup_time + cleanup_time
                max_hours = self._get_max_working_hours(time_slot)
                utilization_rate = (total_time / max_hours * 100) if max_hours > 0 else 0
                
                print(f"    {time_slot}:")
                for prod in productions:
                    product_name = self._get_product_name(prod['product'])
                    print(f"      - {prod['product']} ({product_name}): {prod['production_time']:.1f}시간 "
                          f"({prod['production_quantity_boxes']:.0f}박스)")
                
                # 시간 요약 정보 (작업 준비 시간과 청소 시간 구분)
                if changeover_time > 0 or setup_time > 0 or cleanup_time > 0:
                    time_components = []
                    time_components.append(f"생산: {total_production_time:.1f}h")
                    if changeover_time > 0:
                        time_components.append(f"교체: {changeover_time:.1f}h")
                    if setup_time > 0:
                        time_components.append(f"준비: {setup_time:.1f}h")
                    if cleanup_time > 0:
                        time_components.append(f"청소: {cleanup_time:.1f}h")
                    
                    time_summary = " + ".join(time_components)
                    print(f"      [시간 분석] {time_summary} = 총 {total_time:.1f}h ({utilization_rate:.1f}% 활용)")
                else:
                    print(f"      [시간 분석] 생산: {total_production_time:.1f}h = 총 {total_time:.1f}h ({utilization_rate:.1f}% 활용)")
        
        # 교체 이벤트 상세 정보
        if solution['changeover_events']:
            print(f"\n🔄 교체 이벤트 상세:")
            print("-" * 80)
            print(f"{'호기':<6} {'시간대':<15} {'이전제품':<12} {'다음제품':<12} {'교체시간':<10} {'교체비율':<10}")
            print("-" * 80)
            
            # 전체 교체시간 계산
            total_changeover = sum(event['changeover_time'] for event in solution['changeover_events'])
            
            # JSON 생성과 동일한 정렬된 교체 이벤트 사용
            sorted_changeover_events = self._format_changeover_events_for_json(solution)
            
            for event in sorted_changeover_events:
                changeover_ratio = (event['changeover_time_hours'] / total_changeover * 100) if total_changeover > 0 else 0
                
                # 제품 정보 가져오기
                from_product = event.get('from_product_code', 'N/A')
                to_product = event.get('to_product_code', 'N/A')
                
                if from_product != 'N/A' and to_product != 'N/A':
                    from_name = event.get('from_product_name', self._get_product_name(from_product))
                    to_name = event.get('to_product_name', self._get_product_name(to_product))
                    print(f"{event['line']:<6} {event['time_slot']:<15} {from_product:<12} {to_product:<12} {event['changeover_time_hours']:<10.1f} {changeover_ratio:<9.1f}%")
                    print(f"{'':6} {'':15} {from_name:<12} {to_name:<12}")
                else:
                    print(f"{event['line']:<6} {event['time_slot']:<15} {'N/A':<12} {'N/A':<12} {event['changeover_time_hours']:<10.1f} {changeover_ratio:<9.1f}%")
        else:
            print(f"\n🔄 교체 이벤트: 없음")
        
        # 청소 이벤트 (작업 준비 시간과 청소 시간 구분)
        if solution['cleaning_events']:
            print(f"\n🧹 작업 준비 및 청소 이벤트:")
            
            # 작업 준비 이벤트
            setup_events = [e for e in solution['cleaning_events'] if e['time_slot'] == self.time_slots[0]]
            if setup_events:
                print(f"  📋 작업 준비:")
                for event in setup_events:
                    print(f"    {event['line']}호기 {event['time_slot']}: {event['cleaning_time']:.1f}시간")
            
            # 청소 이벤트
            cleanup_events = [e for e in solution['cleaning_events'] if e['time_slot'] != self.time_slots[0]]
            if cleanup_events:
                print(f"  🧹 청소:")
                for event in cleanup_events:
                    print(f"    {event['line']}호기 {event['time_slot']}: {event['cleaning_time']:.1f}시간")
        
        print("="*80)
    
    def _get_boxes_per_hour(self, product: str, line: str) -> float:
        """
        박스/시간 계산 (트랙 수 고려)
        """
        capacity = self._get_capacity(product, line)  # 시간당 개수 (트랙 수 포함)
        package_count = self._get_package_count(product)
        if package_count > 0:
            return capacity / package_count
        return 0
    
    def _get_ct_rate(self, product: str, line: str) -> float:
        """
        C/T Rate 가져오기 (분당 생산 개수)
        """
        return self._get_capacity_rate(product, line)
    
    def _get_capacity(self, product: str, line: str) -> float:
        """
        Capa 가져오기 (시간당 생산량) - 트랙 수 고려
        """
        ct_rate = self._get_capacity_rate(product, line)  # 분당 개수
        track_count = self._get_track_count(line)  # 트랙 수
        return ct_rate * track_count * 60  # 분당 → 시간당 변환 (트랙 수 포함)
    
    def _get_changeover_type(self, from_product: str, to_product: str) -> str:
        """교체 유형 판단 (일반교체 또는 청소교체)"""
        # 청소가 필요한 교체인지 확인하는 로직
        # 실제 구현에서는 제품 카테고리나 특성에 따라 판단
        
        # 기본 라인 사용 (첫 번째 활성 라인)
        default_line = self.lines[0] if self.lines else '16'
        changeover_time = self._get_changeover_time(from_product, to_product, default_line)
        
        # 교체시간이 긴 경우 청소교체로 판단 (임계값은 조정 가능)
        if changeover_time > 2.0:  # 2시간 이상이면 청소교체
            return "청소교체"
        else:
            return "일반교체"
    
    def _get_product_info_for_json(self) -> Dict:
        """제품 정보를 JSON용으로 정리"""
        product_info = {}
        
        for product_code in self.products:
            product_info[product_code] = {
                "name": self._get_product_name(product_code),
                "package_count": self._get_package_count(product_code),
                "available_lines": [line for line in self.lines if (product_code, line) in self.valid_product_line_combinations]
            }
        
        return product_info
    
    def _get_line_info_for_json(self) -> Dict:
        """라인 정보를 JSON용으로 정리"""
        line_info = {}
        
        for line in self.lines:
            line_info[line] = {
                "track_count": self._get_track_count(line),
                "available_products": [product for product in self.products if (product, line) in self.valid_product_line_combinations]
            }
        
        return line_info
    
    def export_to_excel(self, solution: Dict, output_path: str = None):
        """
        최적화 결과를 엑셀 표 형태로 내보내기
        Args:
            solution: dict, 최적화 결과
            output_path: str, 출력 파일 경로 (None이면 자동 생성)
        """
        if not solution:
            self.logger.error("출력할 결과가 없습니다")
            return
        
        self.logger.info("엑셀 파일 생성 중...")
        
        # 출력 파일 경로 설정
        if output_path is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"production_schedule_{timestamp}.xlsx"
        
        # 엑셀 파일 생성
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # 1. 제품별 달성율 요약 시트
            self._create_achievement_summary_sheet(writer, solution)
            
            # 2. 생산 스케줄 시트
            self._create_production_schedule_sheet(writer, solution)
            
            # 3. 교체/청소 이벤트 시트
            self._create_events_sheet(writer, solution)
        
        self.logger.info(f"엑셀 파일 생성 완료: {output_path}")
        return output_path
    
    def export_to_json(self, solution: Dict, output_path: str = None):
        """
        최적화 결과를 JSON 파일로 내보내기 (상세 정보 포함)
        Args:
            solution: dict, 최적화 결과
            output_path: str, 출력 파일 경로 (None이면 자동 생성)
        """
        import json
        from datetime import datetime
        
        if not solution:
            self.logger.error("출력할 결과가 없습니다")
            return
        
        self.logger.info("JSON 파일 생성 중...")
        
        # 출력 파일 경로 설정
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"production_schedule_detail_{timestamp}.json"
        
        # 상세 정보를 포함한 JSON 데이터 생성
        json_data = {
            "metadata": {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "objective_value": solution.get('objective_value', 0),
                "total_lines": len(self.lines),
                "total_products": len(self.products),
                "time_slots": self.time_slots
            },
            "production_schedule": self._format_production_schedule_for_json(solution),
            "daily_schedule": self._generate_daily_schedule(solution),
            "changeover_events": self._format_changeover_events_for_json(solution),
            "cleaning_events": solution.get('cleaning_events', []),
            "statistics": self._generate_detailed_statistics(solution),
            "product_info": self._get_product_info_for_json(),
            "line_info": self._get_line_info_for_json()
        }
        
        # JSON 파일로 저장
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"JSON 파일 생성 완료: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"JSON 파일 생성 중 오류 발생: {e}")
            return None
    
    def _format_production_schedule_for_json(self, solution: Dict) -> Dict:
        """생산 스케줄을 JSON 형태로 포맷팅"""
        formatted_schedule = {}
        
        for line, schedule in solution['production_schedule'].items():
            formatted_schedule[line] = {}
            
            for time_slot, productions in schedule.items():
                formatted_schedule[line][time_slot] = []
                
                for i, prod in enumerate(productions):
                    # 제품 정보 추가
                    product_name = self._get_product_name(prod['product'])
                    
                    formatted_prod = {
                        "sequence_order": i + 1,  # 생산 순서
                        "product_code": prod['product'],
                        "product_name": product_name,
                        "production_time_hours": round(prod['production_time'], 2),
                        "production_quantity_units": int(prod['production_quantity_units']),
                        "production_quantity_boxes": round(prod['production_quantity_boxes'], 1),
                        "capacity_rate": self._get_capacity_rate(prod['product'], line),
                        "track_count": self._get_track_count(line),
                        "products_per_box": self._get_package_count(prod['product'])
                    }
                    
                    formatted_schedule[line][time_slot].append(formatted_prod)
        
        return formatted_schedule
    
    def _generate_daily_schedule(self, solution: Dict) -> Dict:
        """요일별 스케줄 생성"""
        daily_schedule = {}
        
        # 요일별로 그룹화
        weekdays = ['월요일', '화요일', '수요일', '목요일', '금요일']
        
        for line, schedule in solution['production_schedule'].items():
            daily_schedule[line] = {}
            
            for day_name in weekdays:
                daily_schedule[line][day_name] = {
                    "조간": [],
                    "야간": []
                }
                
                # 해당 요일의 조간/야간 스케줄 추출
                morning_slot = f"{day_name}_조간"
                night_slot = f"{day_name}_야간"
                
                # 스케줄에서 해당 시간대 데이터 찾아서 복사
                if morning_slot in schedule:
                    daily_schedule[line][day_name]["조간"] = schedule[morning_slot].copy()
                
                if night_slot in schedule:
                    daily_schedule[line][day_name]["야간"] = schedule[night_slot].copy()
        
        return daily_schedule
    
    def _format_changeover_events_for_json(self, solution: Dict) -> List[Dict]:
        """교체 이벤트를 JSON 형태로 포맷팅 (시간대 순서 + 교체 발생 순서로 정렬)"""
        formatted_events = []
        
        # 교체 이벤트를 시간대 순서와 교체 발생 순서로 정렬하기 위한 정보 수집
        changeover_events_with_order = []
        
        for event in solution.get('changeover_events', []):
            # 제품 정보가 있는 경우와 없는 경우 처리
            from_product = event.get('from_product')
            to_product = event.get('to_product')
            
            # 시간대와 제품 순서 정보 추출
            time_slot = event['time_slot']
            line = event['line']
            
            # 시간대 순서 인덱스 계산
            time_order = self._get_time_slot_order(time_slot)
            
            # 같은 시간대 내에서의 교체 순서 계산
            # 교체 이벤트가 발생하는 순서대로 정렬 (from_product의 생산 순서 기준)
            changeover_order = self._get_changeover_order_in_timeslot(solution, line, time_slot, from_product, to_product)
            
            changeover_events_with_order.append({
                'event': event,
                'time_order': time_order,
                'changeover_order': changeover_order,
                'line': line,
                'time_slot': time_slot
            })
        
        # 정렬: 1) 시간대 순서, 2) 교체 발생 순서
        changeover_events_with_order.sort(key=lambda x: (x['time_order'], x['changeover_order']))
        
        # 정렬된 순서대로 포맷팅
        for item in changeover_events_with_order:
            event = item['event']
            from_product = event.get('from_product')
            to_product = event.get('to_product')
            
            formatted_event = {
                "line": event['line'],
                "time_slot": event['time_slot'],
                "changeover_time_hours": round(event['changeover_time'], 2)
            }
            
            # 제품 정보가 있는 경우에만 추가
            if from_product and to_product:
                formatted_event.update({
                    "from_product_code": from_product,
                    "from_product_name": self._get_product_name(from_product),
                    "to_product_code": to_product,
                    "to_product_name": self._get_product_name(to_product),
                    "changeover_type": self._get_changeover_type(from_product, to_product)
                })
            else:
                # 제품 정보가 없는 경우 (일반적인 교체시간만 있는 경우)
                formatted_event.update({
                    "from_product_code": "N/A",
                    "from_product_name": "N/A",
                    "to_product_code": "N/A", 
                    "to_product_name": "N/A",
                    "changeover_type": "일반교체"
                })
            
            formatted_events.append(formatted_event)
        
        return formatted_events
    
    def _get_time_slot_order(self, time_slot: str) -> int:
        """시간대 순서 인덱스 반환 (월요일_조간=0, 월요일_야간=1, ...)"""
        weekdays = ['월요일', '화요일', '수요일', '목요일', '금요일']
        shifts = ['조간', '야간']
        
        day, shift = time_slot.split('_')
        day_index = weekdays.index(day)
        shift_index = shifts.index(shift)
        
        return day_index * 2 + shift_index
    
    def _get_changeover_order_in_timeslot(self, solution: Dict, line: str, time_slot: str, from_product: str, to_product: str) -> int:
        """같은 시간대 내에서 교체 이벤트의 발생 순서 인덱스 반환"""
        productions = solution['production_schedule'].get(line, {}).get(time_slot, [])
        
        # 1. 해당 시간대에서 from_product가 생산되는지 확인
        for i, prod in enumerate(productions):
            if prod['product'] == from_product:
                # 같은 시간대 내 교체: from_product의 생산 위치로 순서 결정
                return i
        
        # 2. from_product가 해당 시간대에 없는 경우 (시간대 간 교체)
        # 이전 시간대에서 from_product가 생산되었는지 확인
        prev_time_slots = self._get_previous_time_slots(time_slot)
        
        for prev_slot in prev_time_slots:
            prev_productions = solution['production_schedule'].get(line, {}).get(prev_slot, [])
            if prev_productions:
                # 이전 시간대의 마지막 제품이 from_product인 경우
                if prev_productions[-1]['product'] == from_product:
                    # 시간대 간 교체는 해당 시간대의 첫 번째 교체로 처리
                    # to_product가 해당 시간대의 몇 번째 제품인지 확인
                    for i, prod in enumerate(productions):
                        if prod['product'] == to_product:
                            # 시간대 간 교체는 -1000 + to_product의 위치로 우선순위 부여
                            # 이렇게 하면 시간대 간 교체가 같은 시간대 내 교체보다 먼저 옴
                            return -1000 + i
        
        # 3. 찾을 수 없는 경우 큰 값 반환 (정렬 시 뒤로)
        return 9999
    
    def _get_previous_time_slots(self, current_time_slot: str) -> List[str]:
        """현재 시간대 이전의 시간대들을 반환"""
        weekdays = ['월요일', '화요일', '수요일', '목요일', '금요일']
        shifts = ['조간', '야간']
        
        day, shift = current_time_slot.split('_')
        day_index = weekdays.index(day)
        shift_index = shifts.index(shift)
        
        previous_slots = []
        
        # 같은 날의 이전 교대
        if shift_index > 0:  # 야간인 경우 조간 추가
            previous_slots.append(f"{day}_조간")
        
        # 이전 날의 마지막 교대
        if day_index > 0:  # 첫째 날이 아닌 경우
            prev_day = weekdays[day_index - 1]
            previous_slots.append(f"{prev_day}_야간")
        
        return previous_slots
    
    def _generate_detailed_statistics(self, solution: Dict) -> Dict:
        """상세 통계 정보 생성"""
        # 추가 통계 계산
        total_production_time = 0
        total_changeover_time = 0
        total_changeover_count = 0
        total_cleaning_time = 0
        line_utilization = {}
        
        # 생산시간 및 라인별 가동률 계산
        for line, schedule in solution['production_schedule'].items():
            line_production_time = 0
            line_changeover_time = 0
            line_cleaning_time = 0
            
            for time_slot, productions in schedule.items():
                for prod in productions:
                    total_production_time += prod['production_time']
                    line_production_time += prod['production_time']
            
            # 교체시간 계산 (해당 라인의 교체 이벤트)
            for event in solution.get('changeover_events', []):
                if event['line'] == line:
                    line_changeover_time += event['changeover_time']
                    total_changeover_time += event['changeover_time']
                    total_changeover_count += 1
            
            # 청소시간 계산 (해당 라인의 청소 이벤트)
            for event in solution.get('cleaning_events', []):
                if event['line'] == line:
                    line_cleaning_time += event['cleaning_time']
                    total_cleaning_time += event['cleaning_time']
            
            # 라인별 가동률 계산 (실제 제약시간 기준)
            total_line_time = line_production_time + line_changeover_time + line_cleaning_time
            
            # 시간대별 최대 가동시간 계산
            max_available_time = 0
            for time_slot in self.time_slots:
                if time_slot in schedule:
                    # 수요일은 8시간, 나머지는 10.5시간
                    max_hours = 8.0 if '수요일' in time_slot else 10.5
                    max_available_time += max_hours
            
            # 가동률 계산 (100% 초과 방지)
            utilization_rate = min((total_line_time / max_available_time * 100) if max_available_time > 0 else 0, 100.0)
            line_utilization[line] = round(utilization_rate, 1)
        
        # 전체 효율성 계산
        total_working_time = total_production_time + total_changeover_time + total_cleaning_time
        overall_efficiency = (total_production_time / total_working_time * 100) if total_working_time > 0 else 0
        
        detailed_stats = {
            "objective_value": solution.get('objective_value', 0),
            "total_production_time_hours": round(total_production_time, 2),
            "total_changeover_time_hours": round(total_changeover_time, 2),
            "total_changeover_count": total_changeover_count,
            "average_changeover_time_hours": round(total_changeover_time / max(total_changeover_count, 1), 2),
            "total_cleaning_time_hours": round(total_cleaning_time, 2),
            "line_utilization_percent": line_utilization,
            "overall_efficiency_percent": round(overall_efficiency, 1),
            "total_production_time": total_production_time,
            "total_changeover_time": total_changeover_time,
            "total_cleaning_time": total_cleaning_time,
            "total_working_time": total_working_time
        }
        
        return detailed_stats
    
    def _create_achievement_summary_sheet(self, writer, solution):
        """제품별 달성율 요약 시트 생성 (박스 단위)"""
        # 제품별 총 생산량 계산 (박스 단위)
        product_total_production = {}
        product_production_lines = {}  # 제품별 생산 호기 정보
        
        for line, schedule in solution['production_schedule'].items():
            for time_slot, productions in schedule.items():
                for prod in productions:
                    product_code = prod['product']
                    if product_code not in product_total_production:
                        product_total_production[product_code] = 0
                        product_production_lines[product_code] = set()
                    
                    product_total_production[product_code] += prod['production_quantity_boxes']
                    product_production_lines[product_code].add(line)
        
        # 요약 데이터 생성
        summary_data = []
        for product_code in sorted(self.products):
            target_quantity = self.order_data.get(product_code, 0)
            actual_quantity = product_total_production.get(product_code, 0)
            achievement_rate = (actual_quantity / target_quantity * 100) if target_quantity > 0 else 0
            product_name = self._get_product_name(product_code)
            
            # 생산 호기 정보
            production_lines = product_production_lines.get(product_code, set())
            lines_str = ", ".join(sorted(production_lines)) if production_lines else "미생산"
            
            # 달성율에 따른 상태 표시 (정확한 비교를 위해 소수점 처리)
            achievement_rate_rounded = round(achievement_rate, 1)
            if achievement_rate_rounded >= 100.0:
                achievement_status = "✅ 달성"
            elif achievement_rate_rounded >= 90.0:
                achievement_status = "⚠️ 부족"
            else:
                achievement_status = "❌ 미달성"
            
            summary_data.append({
                '제품코드': product_code,
                '제품명': product_name,
                '목표수량(박스)': target_quantity,
                '실제생산량(박스)': actual_quantity,
                '달성율(%)': achievement_rate_rounded,
                '생산호기': lines_str,
                '달성상태': achievement_status
            })
        
        # 전체 요약 추가
        total_target = sum(self.order_data.values())
        total_actual = sum(product_total_production.values())
        total_achievement = (total_actual / total_target * 100) if total_target > 0 else 0
        total_achievement_rounded = round(total_achievement, 1)
        
        if total_achievement_rounded >= 100.0:
            total_achievement_status = "✅ 달성"
        elif total_achievement_rounded >= 90.0:
            total_achievement_status = "⚠️ 부족"
        else:
            total_achievement_status = "❌ 미달성"
        
        # 교체시간 통계 추가
        total_changeover_time = sum(event['changeover_time'] for event in solution['changeover_events'])
        total_changeover_count = len(solution['changeover_events'])
        avg_changeover_time = total_changeover_time / total_changeover_count if total_changeover_count > 0 else 0
        
        summary_data.append({
            '제품코드': 'TOTAL',
            '제품명': '전체',
            '목표수량(박스)': total_target,
            '실제생산량(박스)': total_actual,
            '달성율(%)': total_achievement_rounded,
            '생산호기': '전체',
            '달성상태': total_achievement_status
        })
        
        # 교체시간 요약 추가
        summary_data.append({
            '제품코드': 'CHANGEOVER',
            '제품명': '교체시간 통계',
            '목표수량(박스)': f'{total_changeover_count}회',
            '실제생산량(박스)': f'{total_changeover_time:.1f}시간',
            '달성율(%)': f'{avg_changeover_time:.1f}시간/회',
            '생산호기': f'총 {total_changeover_count}회',
            '달성상태': f'평균 {avg_changeover_time:.1f}시간'
        })
        
        # 청소시간 요약 추가
        total_cleaning_time = sum(event['cleaning_time'] for event in solution.get('cleaning_events', []))
        total_cleaning_count = len(solution.get('cleaning_events', []))
        avg_cleaning_time = total_cleaning_time / total_cleaning_count if total_cleaning_count > 0 else 0
        
        summary_data.append({
            '제품코드': 'CLEANING',
            '제품명': '청소시간 통계',
            '목표수량(박스)': f'{total_cleaning_count}회',
            '실제생산량(박스)': f'{total_cleaning_time:.1f}시간',
            '달성율(%)': f'{avg_cleaning_time:.1f}시간/회',
            '생산호기': f'총 {total_cleaning_count}회',
            '달성상태': f'평균 {avg_cleaning_time:.1f}시간'
        })
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='달성율_요약', index=False)
        
        # 스타일 적용
        worksheet = writer.sheets['달성율_요약']
        self._apply_excel_styling(worksheet, df_summary)
    
    def _create_production_schedule_sheet(self, writer, solution):
        """생산 스케줄 시트 생성 (박스 단위 포함)"""
        # 데이터 준비
        excel_data = []
        
        # 각 라인별로 데이터 수집
        for line in self.lines:
            line_schedule = solution['production_schedule'].get(line, {})
            
            # 해당 라인에서 생산된 제품들 수집
            line_products = set()
            for time_slot, productions in line_schedule.items():
                for prod in productions:
                    line_products.add(prod['product'])
            
            # 각 제품별로 행 생성
            for product in line_products:
                row_data = {
                    '호기': line,
                    '제품코드': product,
                    '제품명': self._get_product_name(product),  # 제품명 가져오기
                    'C/T': self._get_ct_rate(product, line),  # C/T Rate
                    '개입': self._get_package_count(product),  # 개입수
                    'Capa': self._get_capacity(product, line),  # Capa
                    '박스/시간': self._get_boxes_per_hour(product, line)  # 박스/시간
                }
                
                # 각 시간대별 생산시간 추가
                for time_slot in self.time_slots:
                    production_time = 0
                    production_boxes = 0
                    if time_slot in line_schedule:
                        for prod in line_schedule[time_slot]:
                            if prod['product'] == product:
                                production_time = prod['production_time']
                                production_boxes = prod['production_quantity_boxes']
                                break
                    
                    # 시간대별 컬럼명 생성
                    day, shift = time_slot.split('_')
                    column_name = f"{day}_{shift}"
                    row_data[column_name] = production_time
                    
                    # 박스 단위 컬럼 추가
                    column_name_boxes = f"{day}_{shift}_박스"
                    row_data[column_name_boxes] = production_boxes
                
                excel_data.append(row_data)
        
        # DataFrame 생성
        df = pd.DataFrame(excel_data)
        
        # 컬럼 순서 정렬 (박스 단위 포함)
        columns = ['호기', '제품코드', '제품명', 'C/T', '개입', 'Capa', '박스/시간']
        for time_slot in self.time_slots:
            day, shift = time_slot.split('_')
            columns.append(f"{day}_{shift}")
            columns.append(f"{day}_{shift}_박스")
        
        df = df[columns]
        
        # 엑셀에 저장
        df.to_excel(writer, sheet_name='생산계획', index=False)
        
        # 스타일 적용
        worksheet = writer.sheets['생산계획']
        self._apply_excel_styling(worksheet, df)
    
    def _create_events_sheet(self, writer, solution):
        """교체/청소 이벤트 시트 생성"""
        events_data = []
        
        # 교체 이벤트
        for event in solution['changeover_events']:
            events_data.append({
                '이벤트유형': '교체',
                '호기': event['line'],
                '시간대': event['time_slot'],
                '소요시간(시간)': event['changeover_time']
            })
        
        # 청소 이벤트
        for event in solution['cleaning_events']:
            events_data.append({
                '이벤트유형': '청소',
                '호기': event['line'],
                '시간대': event['time_slot'],
                '소요시간(시간)': event['cleaning_time']
            })
        
        if events_data:
            df_events = pd.DataFrame(events_data)
            df_events.to_excel(writer, sheet_name='교체청소_이벤트', index=False)
            
            # 스타일 적용
            worksheet = writer.sheets['교체청소_이벤트']
            self._apply_excel_styling(worksheet, df_events)
        
        # 교체시간 분석 시트 추가
        self._create_changeover_analysis_sheet(writer, solution)
    
    def _create_changeover_analysis_sheet(self, writer, solution):
        """교체시간 분석 시트 생성"""
        # 라인별 교체시간 분석
        line_changeover_times = {}
        for event in solution['changeover_events']:
            line = event['line']
            if line not in line_changeover_times:
                line_changeover_times[line] = 0
            line_changeover_times[line] += event['changeover_time']
        
        line_analysis_data = []
        for line in sorted(line_changeover_times.keys()):
            total_time = line_changeover_times[line]
            count = len([e for e in solution['changeover_events'] if e['line'] == line])
            avg_time = total_time / count if count > 0 else 0
            line_analysis_data.append({
                '호기': line,
                '총 교체시간(시간)': total_time,
                '교체횟수': count,
                '평균교체시간(시간)': avg_time
            })
        
        if line_analysis_data:
            df_line_analysis = pd.DataFrame(line_analysis_data)
            df_line_analysis.to_excel(writer, sheet_name='교체시간_분석', index=False)
            
            # 스타일 적용
            worksheet = writer.sheets['교체시간_분석']
            self._apply_excel_styling(worksheet, df_line_analysis)
        
        # 시간대별 교체시간 분석
        time_slot_changeover = {}
        for event in solution['changeover_events']:
            time_slot = event['time_slot']
            if time_slot not in time_slot_changeover:
                time_slot_changeover[time_slot] = 0
            time_slot_changeover[time_slot] += event['changeover_time']
        
        time_analysis_data = []
        for time_slot in sorted(time_slot_changeover.keys()):
            total_time = time_slot_changeover[time_slot]
            count = len([e for e in solution['changeover_events'] if e['time_slot'] == time_slot])
            time_analysis_data.append({
                '시간대': time_slot,
                '총 교체시간(시간)': total_time,
                '교체횟수': count,
                '평균교체시간(시간)': total_time / count if count > 0 else 0
            })
        
        if time_analysis_data:
            df_time_analysis = pd.DataFrame(time_analysis_data)
            df_time_analysis.to_excel(writer, sheet_name='시간대별_교체시간', index=False)
            
            # 스타일 적용
            worksheet = writer.sheets['시간대별_교체시간']
            self._apply_excel_styling(worksheet, df_time_analysis)
    
    def _apply_excel_styling(self, worksheet, df):
        """
        엑셀 스타일 적용
        """
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 스타일 적용
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 데이터 셀 스타일 적용
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # 숫자 컬럼은 우측 정렬
                if col > 3:  # C/T, 개입, Capa, 박스/시간, 시간대별 컬럼들
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # 컬럼 너비 자동 조정
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _print_product_analysis(self):
        """
        제품별 목표 생산량, 필요 생산시간, 블록 개수 분석 출력
        """
        print(f"\n📋 제품별 목표 생산량 및 블록 분석:")
        print("-" * 100)
        print(f"{'제품코드':<15} {'제품명':<25} {'목표량(박스)':<12} {'필요시간(h)':<12} {'블록개수':<8} {'분류':<8}")
        print("-" * 100)
        
        # 제품별 분석 데이터 수집
        product_analysis = []
        
        for product in self.products:
            # 목표 생산량
            target_boxes = self.order_data.get(product, 0)
            if target_boxes <= 0:
                continue
                
            # 제품명
            product_name = self._get_product_name(product)
            
            # 가장 적합한 라인 찾기 (첫 번째 유효한 조합 사용)
            best_line = None
            for line in self.lines:
                if (product, line) in self.valid_product_line_combinations:
                    best_line = line
                    break
            
            if not best_line:
                continue
                
            # 필요 생산시간 계산
            required_hours = self._calculate_required_production_hours(product, best_line)
            
            # 블록 개수 계산 (optimizer의 메서드 사용)
            required_blocks = self.optimizer._calculate_required_time_slots(product, best_line)
            
            # 소량/대량 분류
            max_hours = self._get_max_working_hours(self.time_slots[0])  # 기준 시간대 사용
            classification = "소량생산" if required_hours < max_hours else "대량생산"
            
            product_analysis.append({
                'product': product,
                'product_name': product_name,
                'target_boxes': target_boxes,
                'required_hours': required_hours,
                'required_blocks': required_blocks,
                'classification': classification
            })
        
        # 필요 시간 순으로 정렬 (대량생산 → 소량생산)
        product_analysis.sort(key=lambda x: x['required_hours'], reverse=True)
        
        # 출력
        for analysis in product_analysis:
            print(f"{analysis['product']:<15} {analysis['product_name']:<25} "
                  f"{analysis['target_boxes']:<12} {analysis['required_hours']:<12.1f} "
                  f"{analysis['required_blocks']:<8} {analysis['classification']:<8}")
        
        print("-" * 100)
        
        # 요약 통계
        total_products = len(product_analysis)
        large_products = len([p for p in product_analysis if p['classification'] == '대량생산'])
        small_products = len([p for p in product_analysis if p['classification'] == '소량생산'])
        total_blocks = sum(p['required_blocks'] for p in product_analysis)
        total_target_boxes = sum(p['target_boxes'] for p in product_analysis)
        
        print(f"📊 요약: 전체 {total_products}개 제품, 대량생산 {large_products}개, 소량생산 {small_products}개")
        print(f"📦 총 목표량: {total_target_boxes:,}박스, 총 블록수: {total_blocks}개")
        
    def _calculate_required_production_hours(self, product: str, line: str) -> float:
        """
        제품별 목표 생산량 달성에 필요한 순수 생산시간 계산
        Args:
            product: str, 제품 코드
            line: str, 라인 ID
        Returns:
            float: 필요한 생산시간 (시간)
        """
        try:
            # 목표 생산량
            target_boxes = self.order_data.get(product, 0)
            if target_boxes <= 0:
                return 0.0
                
            # 생산 능력 계산
            capacity_rate = self._get_capacity_rate(product, line)  # 분당 생산 개수
            track_count = self._get_track_count(line)  # 트랙 수
            products_per_box = self._get_package_count(product)  # 개입수
            
            if capacity_rate <= 0 or products_per_box <= 0:
                return 0.0
                
            # 시간당 박스 생산량 = (분당 개수 * 트랙 수 * 60분) / 개입수
            boxes_per_hour = (capacity_rate * track_count * 60) / products_per_box
            
            # 필요한 생산시간 = 목표 박스 / 시간당 박스 생산량
            required_hours = target_boxes / boxes_per_hour
            
            return required_hours
            
        except Exception as e:
            self.logger.warning(f"생산시간 계산 실패 (제품: {product}, 라인: {line}): {e}")
            return 0.0

